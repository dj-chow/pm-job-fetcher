#!/usr/bin/env python3
"""
LinkedIn Networking Tool — Find connections at target companies for PM job search.

Uses Playwright persistent context for LinkedIn session management.
Searches 1st/2nd-degree connections, scores contacts, generates personalized
messages via Claude CLI, and tracks everything in Excel + markdown.

Usage:
    python linkedin_networking.py --login          # First run: manual login
    python linkedin_networking.py                   # Research all companies (5 per session)
    python linkedin_networking.py --company BMO TD  # Specific companies only
    python linkedin_networking.py --test            # Dry run (no Claude messages)
"""

import argparse
import json
import os
import random
import re
import subprocess
import sys
import tempfile
import time
import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_FILE = os.path.join(SCRIPT_DIR, "settings.json")
NETWORKING_COMPANIES_FILE = os.path.join(SCRIPT_DIR, "networking_companies.json")
SESSION_DIR = os.path.join(SCRIPT_DIR, "linkedin_session")
SUMMARY_FILE = os.path.join(SCRIPT_DIR, "networking_summary.md")
CLAUDE_CMD = ["claude.cmd"] if os.name == "nt" else ["claude"]

MAX_SEARCHES_PER_SESSION = 15
COMPANIES_PER_SESSION = 5


def _safe_print(text):
    """Print with ASCII fallback so Windows cp1252 consoles don't crash."""
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="replace").decode("ascii"))


def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE, "r") as f:
            return json.load(f)
    return {}


def load_networking_companies():
    """Load flat list of company names for LinkedIn networking."""
    with open(NETWORKING_COMPANIES_FILE, "r") as f:
        return json.load(f)


# ── Session Management ────────────────────────────────────────────────────────

def login_flow():
    """Open a visible browser for manual LinkedIn login. Saves session cookies."""
    from playwright.sync_api import sync_playwright

    os.makedirs(SESSION_DIR, exist_ok=True)
    print("Opening browser for LinkedIn login...")
    print("Log in manually, then close the browser when done.\n")

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            SESSION_DIR,
            headless=False,
            viewport={"width": 1280, "height": 800},
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto("https://www.linkedin.com/login", wait_until="domcontentloaded")

        # Wait for user to log in and close the browser
        try:
            page.wait_for_event("close", timeout=300_000)  # 5 min max
        except Exception:
            pass

        try:
            context.close()
        except Exception:
            pass  # Browser already closed by user — session is still saved

    print("Session saved to linkedin_session/")
    print("You can now run: python linkedin_networking.py --company BMO --test")


def create_browser_context(playwright, headless=True):
    """Create a persistent browser context with saved LinkedIn session."""
    if not os.path.exists(SESSION_DIR):
        print("ERROR: No saved session. Run with --login first.")
        sys.exit(1)

    context = playwright.chromium.launch_persistent_context(
        SESSION_DIR,
        headless=headless,
        viewport={"width": 1280, "height": 800},
    )
    return context


# ── Connection Discovery ──────────────────────────────────────────────────────

def _random_delay(low=3, high=8):
    """Rate limiting between LinkedIn actions."""
    time.sleep(random.uniform(low, high))


def _check_for_challenge(page):
    """Return True if LinkedIn is showing a challenge/captcha page."""
    url = page.url.lower()
    if "checkpoint" in url or "challenge" in url or "authwall" in url:
        return True
    # Check for common challenge text
    try:
        content = page.content()
        if "security verification" in content.lower():
            return True
    except Exception:
        pass
    return False


def search_connections(page, company, network_degree, search_query=None, search_count=None):
    """
    Search LinkedIn for connections at a company.

    Args:
        page: Playwright page
        company: Company name
        network_degree: "F" for 1st degree, "S" for 2nd degree
        search_query: Additional search terms (e.g., "product manager")
        search_count: Mutable list [count] to track total searches this session
    """
    if search_count and search_count[0] >= MAX_SEARCHES_PER_SESSION:
        _safe_print(f"    Hit max searches ({MAX_SEARCHES_PER_SESSION}), skipping")
        return []

    keywords = company
    if search_query:
        keywords = f"{company} {search_query}"

    # URL-encode the keywords
    import urllib.parse
    encoded_keywords = urllib.parse.quote(keywords)
    network_param = urllib.parse.quote(f'["{network_degree}"]')

    url = (f"https://www.linkedin.com/search/results/people/"
           f"?keywords={encoded_keywords}&network={network_param}")

    degree_label = "1st" if network_degree == "F" else "2nd"
    query_label = f" ({search_query})" if search_query else ""
    _safe_print(f"    Searching {degree_label}-degree connections{query_label}...")

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30_000)
        _random_delay(2, 4)

        if _check_for_challenge(page):
            _safe_print("    WARNING: LinkedIn challenge page detected — stopping searches")
            if search_count:
                search_count[0] = MAX_SEARCHES_PER_SESSION  # prevent further searches
            return []

        if search_count:
            search_count[0] += 1

        # Wait for results to load — try multiple known selectors
        # LinkedIn uses data-view-name on result title links
        try:
            page.wait_for_selector(
                "a[data-view-name='search-result-lockup-title']", timeout=8_000
            )
        except Exception:
            pass

        # Check if any results exist
        result_links = page.query_selector_all(
            "a[data-view-name='search-result-lockup-title']"
        )
        if not result_links:
            # Debug: save screenshot + page snippet to help fix selectors
            debug_dir = os.path.join(SCRIPT_DIR, "debug")
            os.makedirs(debug_dir, exist_ok=True)
            try:
                page.screenshot(path=os.path.join(debug_dir, "search_results.png"))
                html = page.content()
                with open(os.path.join(debug_dir, "search_results.html"), "w",
                          encoding="utf-8") as f:
                    f.write(html)
                _safe_print(f"    No results found (debug saved to debug/)")
            except Exception:
                _safe_print("    No results found")
            return []

        # Extract contacts from search results
        contacts = _extract_contacts(page, company, network_degree)
        _safe_print(f"    Found {len(contacts)} contacts")
        return contacts

    except Exception as e:
        _safe_print(f"    Search error: {e}")
        return []


def _extract_contacts(page, company, network_degree):
    """Extract contact info from LinkedIn search results page."""
    contacts = []

    try:
        # Each result has a title link with data-view-name="search-result-lockup-title"
        # Walk up to the listitem container to get headline/mutual info
        title_links = page.query_selector_all(
            "a[data-view-name='search-result-lockup-title']"
        )

        for link_el in title_links[:10]:  # Cap at 10 per search
            try:
                name = link_el.inner_text().strip()
                if not name or "linkedin member" in name.lower():
                    continue

                # Profile URL
                href = link_el.get_attribute("href") or ""
                profile_url = ""
                if "/in/" in href:
                    profile_url = href.split("?")[0]
                    if not profile_url.startswith("http"):
                        profile_url = "https://www.linkedin.com" + profile_url

                # Navigate up to the listitem container to find headline
                # The container is a div[role='listitem'] ancestor
                container = link_el.evaluate_handle(
                    """el => {
                        let node = el;
                        for (let i = 0; i < 10; i++) {
                            node = node.parentElement;
                            if (!node) return null;
                            if (node.getAttribute('role') === 'listitem') return node;
                        }
                        return null;
                    }"""
                )

                headline = ""
                mutual_count = 0

                if container:
                    # Get all text content and parse
                    all_text = container.evaluate("el => el.innerText") or ""
                    lines = [l.strip() for l in all_text.split("\n") if l.strip()]

                    # Headline is typically the line right after the name
                    # (skipping degree indicators like "2nd")
                    found_name = False
                    for line in lines:
                        if name in line:
                            found_name = True
                            continue
                        if found_name and line not in ("", "2nd", "1st", "3rd"):
                            # Skip short labels like degree indicators
                            if len(line) > 5 and not line.startswith("Connect"):
                                headline = line
                                break

                    # Mutual connections
                    for line in lines:
                        m = re.search(r"(\d+)\s*(?:other\s+)?mutual\s+connection", line)
                        if m:
                            mutual_count = int(m.group(1))
                            break
                        # Also catch "Name, Name and 3 other mutual connections"
                        m = re.search(r"and\s+(\d+)\s+other", line)
                        if m:
                            mutual_count = int(m.group(1))
                            break

                degree = "1st" if network_degree == "F" else "2nd"

                contacts.append({
                    "name": name,
                    "headline": headline,
                    "profile_url": profile_url,
                    "degree": degree,
                    "company": company,
                    "mutual_connections": mutual_count,
                })
            except Exception:
                continue
    except Exception as e:
        _safe_print(f"    Extraction error: {e}")

    return contacts


def discover_connections(page, company, search_count):
    """
    Run all searches for a single company.
    Returns a deduplicated list of contacts.
    """
    all_contacts = []

    # 1st-degree connections at this company
    contacts_1st = search_connections(page, company, "F", search_count=search_count)
    all_contacts.extend(contacts_1st)
    _random_delay()

    # 2nd-degree PM-relevant people at this company
    contacts_2nd = search_connections(
        page, company, "S", search_query="product manager", search_count=search_count
    )
    all_contacts.extend(contacts_2nd)
    _random_delay()

    # 2nd-degree recruiters at this company
    contacts_recruit = search_connections(
        page, company, "S", search_query="recruiter talent", search_count=search_count
    )
    all_contacts.extend(contacts_recruit)

    # Dedup by name
    seen_names = set()
    unique = []
    for c in all_contacts:
        key = c["name"].lower()
        if key not in seen_names:
            seen_names.add(key)
            unique.append(c)

    return unique


# ── Contact Scoring ───────────────────────────────────────────────────────────

PM_TITLE_PATTERNS = [
    "product manager", "product lead", "product owner", "group product",
    "head of product", "director of product", "vp product", "pm ",
    "technical program manager", "tpm",
]

RECRUITER_PATTERNS = [
    "recruiter", "talent acquisition", "talent partner",
    "people operations", "hiring",
]

LEADERSHIP_PATTERNS = [
    "director", "vp ", "vice president", "head of", "svp", "evp",
    "chief", "managing director",
]


def score_contact(contact):
    """Score a contact based on degree, role relevance, and mutual connections."""
    score = 0
    headline_lower = contact.get("headline", "").lower()

    # Connection degree
    if contact["degree"] == "1st":
        score += 20
    elif contact["degree"] == "2nd":
        score += 10

    # Role relevance
    if any(p in headline_lower for p in PM_TITLE_PATTERNS):
        score += 15
    elif any(p in headline_lower for p in RECRUITER_PATTERNS):
        score += 12
    elif any(p in headline_lower for p in LEADERSHIP_PATTERNS):
        score += 8

    # Mutual connections (2 pts each, max 10)
    mutual = contact.get("mutual_connections", 0)
    score += min(mutual * 2, 10)

    return score


def rank_contacts(contacts):
    """Score and sort contacts by relevance."""
    for c in contacts:
        c["score"] = score_contact(c)
    contacts.sort(key=lambda c: c["score"], reverse=True)
    return contacts


# ── Message Generation via Claude CLI ─────────────────────────────────────────

def generate_message(contact, test_mode=False):
    """Generate a personalized outreach message using Claude CLI."""
    settings = load_settings()
    user_name = settings.get("user_name", "a PM")
    user_title = settings.get("user_title", "a Product Manager")
    user_experience = settings.get("user_experience", "experience in fintech")
    if test_mode:
        return "[TEST MODE — message would be generated here]"

    company = contact["company"]
    name = contact["name"].split()[0]  # First name
    headline = contact["headline"]
    degree = contact["degree"]

    if degree == "1st":
        prompt = f"""Write a short LinkedIn message (under 100 words) to {name} who works at {company}.
Their headline: "{headline}"

Context: I'm {user_name}, {user_title} with {user_experience}.
I'm exploring Product Manager roles at {company}. We're already connected on LinkedIn.

Tone: warm, specific, respectful. Ask if they'd be open to a quick chat about the PM team.
Do NOT use phrases like "I noticed" or "I came across". Be direct and genuine.
Output ONLY the message text, no subject line or commentary."""
    else:
        prompt = f"""Write a LinkedIn connection request note (under 300 characters) to {name} at {company}.
Their headline: "{headline}"

Context: I'm {user_name}, {user_title} with {user_experience}.
Interested in PM roles at {company}. We're not connected yet (2nd degree).

Tone: brief, specific, respectful. Mention something relevant to their role.
Output ONLY the note text, no commentary."""

    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".txt", delete=False, encoding="utf-8", errors="replace"
    ) as f:
        f.write(prompt)
        tmpfile = f.name

    try:
        env = os.environ.copy()
        env.pop("CLAUDECODE", None)

        with open(tmpfile, "r", encoding="utf-8", errors="replace") as fin:
            result = subprocess.run(
                CLAUDE_CMD + ["-p", "-"],
                stdin=fin,
                capture_output=True,
                timeout=120,
                env=env,
                encoding="utf-8",
                errors="replace",
            )
    except Exception as e:
        _safe_print(f"    Claude CLI exception: {e}")
        return ""
    finally:
        try:
            os.unlink(tmpfile)
        except Exception:
            pass

    if not result or result.returncode != 0:
        stderr = getattr(result, "stderr", "") or ""
        _safe_print(f"    Claude CLI error: {stderr[:200]}")
        return ""

    return (result.stdout or "").strip()


# ── Excel Tracker — Networking Tab ────────────────────────────────────────────

def update_excel_tracker(contacts, settings):
    """Add contacts to the Networking tab of the Excel tracker."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Alignment, Font
    except ImportError:
        _safe_print("  Skipping Excel update (openpyxl not installed)")
        return

    excel_path = settings.get("excel_tracker_path", "")
    if not os.path.exists(excel_path):
        _safe_print(f"  Excel tracker not found: {excel_path}")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)

        # Create Networking tab if it doesn't exist
        if "Networking" not in wb.sheetnames:
            ws = wb.create_sheet("Networking")
            headers = [
                "Date", "Name", "Company", "Role/Headline", "Degree",
                "Message", "Status", "Follow-Up Date", "Notes", "Profile URL",
            ]
            header_font = Font(bold=True)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
            # Set column widths
            widths = [12, 22, 18, 35, 8, 50, 12, 14, 30, 45]
            for col, width in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width
        else:
            ws = wb["Networking"]

        # Find next empty row
        next_row = 2
        for row in ws.iter_rows(min_row=2, max_col=2):
            if any(cell.value for cell in row):
                next_row = row[0].row + 1

        # Collect existing (name, company) pairs to dedup
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2]:
                existing.add((str(row[1]).strip().lower(), str(row[2]).strip().lower()))

        yellow_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        today_str = datetime.date.today().strftime("%Y-%m-%d")

        added = 0
        for contact in contacts:
            key = (contact["name"].lower(), contact["company"].lower())
            if key in existing:
                continue

            row_data = [
                today_str,                          # A: Date
                contact["name"],                    # B: Name
                contact["company"],                 # C: Company
                contact.get("headline", ""),         # D: Role/Headline
                contact["degree"],                   # E: Degree
                contact.get("message", ""),          # F: Message
                "Pending",                           # G: Status
                "",                                  # H: Follow-Up Date
                f"Score: {contact.get('score', 0)}", # I: Notes
                contact.get("profile_url", ""),      # J: Profile URL
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col, value=value)
                cell.fill = yellow_fill
                cell.alignment = Alignment(wrap_text=(col == 6))  # Wrap message col

            existing.add(key)
            next_row += 1
            added += 1

        wb.save(excel_path)
        _safe_print(f"  Excel tracker updated: {added} new contact(s) in Networking tab")

    except Exception as e:
        _safe_print(f"  Excel update failed: {e}")


# ── Markdown Summary ──────────────────────────────────────────────────────────

def write_summary(all_results):
    """Write networking_summary.md with per-company contact lists and messages."""
    today = datetime.date.today().strftime("%B %d, %Y")
    lines = [
        f"# LinkedIn Networking Summary — {today}\n",
        f"Generated by linkedin_networking.py\n",
    ]

    total_contacts = sum(len(contacts) for contacts in all_results.values())
    lines.append(f"> {total_contacts} contacts across {len(all_results)} companies\n")

    for company, contacts in sorted(all_results.items()):
        lines.append(f"\n## {company}\n")

        first_degree = [c for c in contacts if c["degree"] == "1st"]
        second_degree = [c for c in contacts if c["degree"] == "2nd"]

        if first_degree:
            lines.append("### 1st-Degree Connections (Referral Opportunities)\n")
            for c in first_degree[:5]:
                lines.append(f"- **{c['name']}** — {c.get('headline', 'N/A')}")
                if c.get("profile_url"):
                    lines.append(f"  - Profile: {c['profile_url']}")
                lines.append(f"  - Score: {c.get('score', 0)}")
                if c.get("message"):
                    lines.append(f"  - Message: {c['message'][:200]}")
                lines.append("")

        if second_degree:
            lines.append("### Top 2nd-Degree Contacts\n")
            for c in second_degree[:5]:
                lines.append(f"- **{c['name']}** — {c.get('headline', 'N/A')}")
                if c.get("profile_url"):
                    lines.append(f"  - Profile: {c['profile_url']}")
                mutuals = c.get("mutual_connections", 0)
                if mutuals:
                    lines.append(f"  - {mutuals} mutual connection(s)")
                lines.append(f"  - Score: {c.get('score', 0)}")
                if c.get("message"):
                    lines.append(f"  - Connection note: {c['message'][:300]}")
                lines.append("")

        if not first_degree and not second_degree:
            lines.append("No connections found.\n")

    lines.append("\n---\n")
    lines.append("**Next steps:** Send messages to Pending contacts, "
                 "prioritize 1st-degree for referrals.\n")

    with open(SUMMARY_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    _safe_print(f"\nSummary written to: networking_summary.md")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="LinkedIn networking tool for PM job search")
    parser.add_argument("--login", action="store_true", help="Open browser for manual LinkedIn login")
    parser.add_argument("--company", nargs="+", help="Specific companies to search (default: all)")
    parser.add_argument("--test", action="store_true", help="Dry run — no Claude message generation")
    parser.add_argument("--visible", action="store_true", help="Run browser in visible mode (for debugging)")
    args = parser.parse_args()

    if args.login:
        login_flow()
        return

    print("=" * 60)
    print("LinkedIn Networking Tool")
    print("=" * 60)

    settings = load_settings()
    company_names = load_networking_companies()

    # Filter to requested companies
    if args.company:
        # Match by case-insensitive substring
        targets = []
        for requested in args.company:
            matched = [n for n in company_names if requested.lower() in n.lower()]
            targets.extend(matched)
        if not targets:
            print(f"No matching companies found for: {args.company}")
            print(f"Available: {', '.join(company_names)}")
            sys.exit(1)
        targets = list(dict.fromkeys(targets))  # dedup preserving order
    else:
        # Default: first N companies per session
        targets = company_names[:COMPANIES_PER_SESSION]

    print(f"\nTarget companies: {', '.join(targets)}")
    if args.test:
        print("[TEST MODE — no Claude messages will be generated]\n")

    # Launch browser
    from playwright.sync_api import sync_playwright

    all_results = {}
    search_count = [0]  # mutable counter

    with sync_playwright() as p:
        context = create_browser_context(p, headless=not args.visible)
        page = context.pages[0] if context.pages else context.new_page()

        # Verify we're logged in
        page.goto("https://www.linkedin.com/feed/", wait_until="domcontentloaded", timeout=30_000)
        _random_delay(2, 3)

        if "login" in page.url.lower() or "authwall" in page.url.lower():
            print("ERROR: Not logged in. Run with --login first.")
            context.close()
            sys.exit(1)

        print("LinkedIn session active.\n")

        for i, company in enumerate(targets):
            if search_count[0] >= MAX_SEARCHES_PER_SESSION:
                _safe_print(f"\nHit max searches ({MAX_SEARCHES_PER_SESSION}) — stopping")
                break

            print(f"[{i+1}/{len(targets)}] {company}")
            print("-" * 40)

            contacts = discover_connections(page, company, search_count)

            if contacts:
                contacts = rank_contacts(contacts)

                # Generate messages for top contacts
                top_contacts = contacts[:5]  # Message top 5 per company
                for j, contact in enumerate(top_contacts):
                    _safe_print(f"  Generating message for {contact['name']}...")
                    contact["message"] = generate_message(contact, test_mode=args.test)
                    if j < len(top_contacts) - 1:
                        _random_delay(1, 3)

                all_results[company] = contacts
                _safe_print(f"  Total: {len(contacts)} contacts, top score: {contacts[0]['score']}")
            else:
                all_results[company] = []
                _safe_print("  No connections found")

            if i < len(targets) - 1:
                _random_delay()

            print()

        context.close()

    # Flatten all contacts for Excel
    all_contacts = []
    for contacts in all_results.values():
        all_contacts.extend(contacts[:10])  # Top 10 per company

    # Update Excel tracker
    if all_contacts:
        print("\nUpdating trackers...")
        update_excel_tracker(all_contacts, settings)

    # Write markdown summary
    write_summary(all_results)

    total = sum(len(c) for c in all_results.values())
    print(f"\nDone! Found {total} contacts across {len(all_results)} companies.")


if __name__ == "__main__":
    main()
